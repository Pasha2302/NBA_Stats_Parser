from pathlib import Path
import os
import re

import openpyxl.worksheet.worksheet
from openpyxl import load_workbook
# from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import range_boundaries

import toolbox as rwf
from copy import copy

MAIN_SHIFT = 12
path_result_home = str(Path("Result", "result_home.xlsx"))
path_result_away = str(Path("Result", "result_away.xlsx"))
path_searching_results = str(Path("Result", "Searching_Results.xlsx"))


def search_input_team(home_team: str, away_team: str) -> dict:
    list_data_home = []
    list_data_away = []
    data_dict = dict()
    data_dict['home_team'] = []
    data_dict['away_team'] = []

    home_teams = load_workbook(path_result_home)
    away_teams = load_workbook(path_result_away)

    sheet_home: openpyxl.worksheet.worksheet.Worksheet = home_teams.active
    sheet_away: openpyxl.worksheet.worksheet.Worksheet = away_teams.active

    row_generator_home = sheet_home.iter_rows(min_row=2, values_only=True)
    row_generator_away = sheet_away.iter_rows(min_row=2, values_only=True)

    for t1, row_h in enumerate(row_generator_home, start=1):
        if home_team in row_h:
            list_data_home.extend(row_h)
            list_data_home.insert(12, [None for _ in range(16)])
            list_data_home.insert(-1, [None, None])
            break
        # print(f"[{t1}] {row_h}")
    print('==' * 40)
    for t2, row_a in enumerate(row_generator_away, start=1):
        if away_team in row_a:
            list_data_away.extend(row_a)
            list_data_away.insert(12, [None for _ in range(16)])
            list_data_away.insert(-1, [None, None])
            break
        # print(f"[{t2}] {row_a}")

    for i in list_data_home:
        if isinstance(i, list):
            data_dict['home_team'].extend(i)
        else:
            data_dict['home_team'].append(i)

    for i in list_data_away:
        if isinstance(i, list):
            data_dict['away_team'].extend(i)
        else:
            data_dict['away_team'].append(i)

    return data_dict


def format_name_team(name_team):
    name_team_f = re.sub(r"Los Angeles Clippers", "LA Clippers", name_team)
    return name_team_f


def get_result_game_five(game_five: str) -> str:
    count_win = 0
    list_result_g = game_five.split(',')
    for r in list_result_g:
        if 'W' == r:
            count_win += 1
    return f'{count_win}/5'


def add_data_to_file(path_file, dt):
    wb = load_workbook('example.xlsx')
    sheet: openpyxl.worksheet.worksheet.Worksheet = wb.active

    for col, data in enumerate(dt['home_team'], start=1):
        sheet.cell(row=3, column=col, value=data)

    for col2, data2 in enumerate(dt['away_team'], start=1):
        sheet.cell(row=6, column=col2, value=data2)

    coeff_team = rwf.download_json_data(path_file='coeff_team.json')
    for rows in sheet.rows:
        for data_team in coeff_team:
            home_team = data_team['home_team']['name']
            away_team = data_team['away_team']['name']
            list_values = [check_team.value for check_team in rows]

            if home_team in list_values or away_team in list_values:
                for cell in rows:
                    if home_team in list_values and '1?' == cell.value and not isinstance(cell, MergedCell):
                        coeff_home = data_team['home_team']['coeff']
                        if coeff_home:
                            coeff_home = float(coeff_home)
                        cell.value = coeff_home
                    if away_team in list_values and '2?' == cell.value and not isinstance(cell, MergedCell):
                        coeff_away = data_team['away_team']['coeff']
                        if coeff_away:
                            coeff_away = float(coeff_away)
                        cell.value = coeff_away

    result_game_teams = rwf.download_json_data(path_file='teams_result_game.json')
    for rows in sheet.rows:
        for data_result_game in result_game_teams:
            home_team = format_name_team(data_result_game['team_1']['name'])
            away_team = format_name_team(data_result_game['team_2']['name'])
            list_values = [check_team.value for check_team in rows]

            if home_team in list_values or away_team in list_values:
                for cell in rows:
                    if home_team in list_values and '1??' == cell.value and not isinstance(cell, MergedCell):
                        result_game_home = data_result_game['team_1']['five_results']
                        if result_game_home:
                            result_game_home = get_result_game_five(result_game_home)
                        cell.value = result_game_home

                    if away_team in list_values and '2??' == cell.value and not isinstance(cell, MergedCell):
                        result_game_away = data_result_game['team_2']['five_results']
                        if result_game_away:
                            result_game_away = get_result_game_five(result_game_away)
                        cell.value = result_game_away

                    if home_team in list_values and '1???' == cell.value and not isinstance(cell, MergedCell):
                        last_game_home = data_result_game['team_1']['five_results'].split(',')[-1]
                        if last_game_home:
                            cell.value = last_game_home

                    if away_team in list_values and '2???' == cell.value and not isinstance(cell, MergedCell):
                        last_game_home = data_result_game['team_2']['five_results'].split(',')[-1]
                        if last_game_home:
                            cell.value = last_game_home

    wb.save(path_file)
    wb.close()


def copy_cell(src_sheet, src_row, src_col,
              tgt_sheet, tgt_row, tgt_col, following_cells):

    cell: Cell | MergedCell = src_sheet.cell(src_row, src_col)
    new_cell: Cell | MergedCell = tgt_sheet.cell(tgt_row, tgt_col)

    new_cell._style = copy(cell._style)
    # new_cell.border = copy(cell.border)

    if isinstance(cell, Cell) and isinstance(new_cell, Cell):
        new_cell.value = cell.value

    if cell.value:
        if '=ABS(' in str(cell.value):
            col = re.search(r"((?<=\()\w*(?=\d))", cell.value).group()
            new_cell.value = f"=ABS({col}{3 + following_cells}-{col}{6 + following_cells})"

        if not isinstance(new_cell, MergedCell) and '=ABS(' not in str(cell.value):
            new_cell.value = cell.value


def start_create_final_data(home_team, away_team):
    if not os.path.isfile(path_searching_results) and os.path.isfile('formula_cell.json'):
        os.remove('formula_cell.json')
    if not os.path.isfile(path_searching_results) and os.path.isfile('test2.xlsx'):
        os.remove('test2.xlsx')
    if not os.path.isfile('test2.xlsx') and os.path.isfile('insert_rows_count.txt'):
        os.remove('insert_rows_count.txt')

    if os.path.isfile('insert_rows_count.txt'):
        insert_rows_count = eval(rwf.download_txt_data(path_file='insert_rows_count.txt'))
    else:
        insert_rows_count = [10, 11, 12]

    if os.path.isfile('formula_cell.json'):
        formula_cell = rwf.download_json_data(path_file='formula_cell.json')
    else:
        formula_cell = {'following_cells': MAIN_SHIFT}

    data_team = search_input_team(home_team, away_team)
    print(data_team)

    if not os.path.isfile(path_searching_results):
        add_data_to_file(path_file=path_searching_results, dt=data_team)
    else:
        add_data_to_file(path_file='test2.xlsx', dt=data_team)

        wb1 = load_workbook(path_searching_results)
        wb2 = load_workbook('test2.xlsx')
        ws1: openpyxl.worksheet.worksheet.Worksheet = wb1.active
        ws2: openpyxl.worksheet.worksheet.Worksheet = wb2.active

        ws1_last_row = ws1.max_row

        for i, row in enumerate(ws2.iter_rows(), 1):
            for cell in row:
                copy_cell(ws2, cell.row, cell.column,
                          ws1, ws1_last_row + i, cell.column,
                          formula_cell['following_cells']
                          )

        ws1.insert_rows(insert_rows_count[0])
        ws1.insert_rows(insert_rows_count[1])
        ws1.insert_rows(insert_rows_count[2])

        for _range in ws2.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            # print(boundaries)
            ws1.merge_cells(start_column=boundaries[0], start_row=boundaries[1] + insert_rows_count[-1],
                            end_column=boundaries[2], end_row=boundaries[3] + insert_rows_count[-1])

        wb1.save(path_searching_results)

        formula_cell['following_cells'] = formula_cell['following_cells'] + MAIN_SHIFT
        rwf.save_json_data(json_data=formula_cell, path_file='formula_cell.json')

        insert_rows_save = []
        insert_rows_save.extend([
            insert_rows_count[0] + MAIN_SHIFT, insert_rows_count[1] + MAIN_SHIFT, insert_rows_count[2] + MAIN_SHIFT
        ])
        rwf.save_txt_data(data_txt=str(insert_rows_save), path_file='insert_rows_count.txt')

        wb1.close()
        wb2.close()


if __name__ == '__main__':
    # info_row_cell()
    pass
