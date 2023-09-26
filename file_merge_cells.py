from pathlib import Path
import os

from openpyxl import load_workbook
import openpyxl.worksheet.worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import range_boundaries
from copy import copy

import toolbox as rwf


def copy_cell(src_sheet, src_row, src_col,
              tgt_sheet, tgt_row, tgt_col,
              formula_desired_cells1, formula_desired_cells2,
              formula_cell_h1, formula_cell_h2, formula_cell_a1, formula_cell_a2):

    cell: Cell | MergedCell = src_sheet.cell(src_row, src_col)
    new_cell: Cell | MergedCell = tgt_sheet.cell(tgt_row, tgt_col)

    if not isinstance(new_cell, MergedCell):
        new_cell.value = cell.value
    if cell.value:
        if '=SUM(' in str(cell.value) and cell.row + 18 == formula_desired_cells1:
            new_cell.value = f"=SUM(T{formula_cell_h1}:T{formula_cell_h2})"

    if cell.value:
        if '=SUM(' in str(cell.value) and cell.row + 18 == formula_desired_cells2:
            new_cell.value = f"=SUM(T{formula_cell_a1}:T{formula_cell_a2})"
    # if cell.has_style and copy_style:
    new_cell.border = copy(cell.border)
    new_cell._style = copy(cell._style)


def main():
    payh_file = str(Path("Result", "Searching_Results.xlsx"))

    wb1 = load_workbook(payh_file)
    sheet1: openpyxl.worksheet.worksheet.Worksheet = wb1.active
    # ------------------------------------------------------------------------
    insert_rows_count = [16, 17, 18]
    if os.path.isfile('formula_cell.json'):
        formula_cell = rwf.download_json_data(path_file='formula_cell.json')
    else:
        formula_cell = {
            'desired_cells': [2, 9],
            'formula_cell_h': [2, 7],
            'formula_cell_a': [9, 14]
        }

    formula_desired_cells1 = formula_cell['desired_cells'][0] + 18
    formula_desired_cells2 = formula_cell['desired_cells'][1] + 18

    formula_cell_h1 = formula_cell['formula_cell_h'][0] + 18
    formula_cell_h2 = formula_cell['formula_cell_h'][1] + 18

    formula_cell_a1 = formula_cell['formula_cell_a'][0] + 18
    formula_cell_a2 = formula_cell['formula_cell_a'][1] + 18

    wb2 = load_workbook('test2.xlsx')
    sheet2: openpyxl.worksheet.worksheet.Worksheet = wb2.active
    sheet1_last_row = sheet1.max_row

    for _range in sheet2.merged_cells.ranges:
        boundaries = range_boundaries(str(_range))
        # print(boundaries)
        sheet1.merge_cells(start_column=boundaries[0], start_row=boundaries[1] + insert_rows_count[-1],
                           end_column=boundaries[2], end_row=boundaries[3] + insert_rows_count[-1])

    for i, row in enumerate(sheet2.iter_rows(), 1):
        for cell in row:
            copy_cell(sheet2, cell.row, cell.column,
                      sheet1, sheet1_last_row + i, cell.column,
                      formula_desired_cells1, formula_desired_cells2,
                      formula_cell_h1, formula_cell_h2, formula_cell_a1, formula_cell_a2
                      )

    formula_cell['desired_cells'] = [formula_desired_cells1, formula_desired_cells2]
    formula_cell['formula_cell_h'] = [formula_cell_h1, formula_cell_h2]
    formula_cell['formula_cell_a'] = [formula_cell_a1, formula_cell_a2]
    rwf.save_json_data(json_data=formula_cell, path_file='formula_cell.json')

    sheet1.insert_rows(insert_rows_count[0])
    sheet1.insert_rows(insert_rows_count[1])
    sheet1.insert_rows(insert_rows_count[2])
    wb1.save(payh_file)

    wb1.close()
    wb2.close()

    cell: MergedCell = sheet1["U2"]
    print(cell.value, type(cell.value), type(cell))
    print(sheet1.max_row)
    print(cell.row, type(cell.row))

    print('--' * 40)
    print(cell.has_style)
    print(cell.style)
    print('>>' * 40)
    print(cell._style)
    print(cell.fill)


if __name__ == '__main__':
    main()
